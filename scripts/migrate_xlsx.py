#!/usr/bin/env python3
"""
Migreert VfY_FotoBewerking_Tracker.xlsx naar SQL voor de Supabase-database.

Gebruik:
    python migrate_xlsx.py VfY_FotoBewerking_Tracker.xlsx -o ../db/03_seed_data.sql

Uitvoer:
    - 03_seed_data.sql : INSERT-statements voor rental_experts, editors,
                         assignments en edit_items
    - migratie_log.txt : elke aanname, conversie en overgeslagen waarde

Aannames staan bovenaan als constanten en worden in het log herhaald.
"""

import argparse
import re
import sys
from collections import Counter, OrderedDict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl ontbreekt. Installeer met: pip install openpyxl")

# ── Configuratie ─────────────────────────────────────────────────────────────

SHEET = "Tracker"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Kolomindex (1-based) uit het Excel-bestand
COL = {
    "acco_id": 1,
    "goals_start": 2,   # B t/m H
    "goals_end": 8,
    "count": 9,
    "assigned_to": 10,
    "date_assigned": 11,
    "status": 12,
    "priority": 13,
    "date_completed": 14,
    "notes_qc": 15,
    "request_date": 16,
    "rental_expert": 17,
}

# Kolomvolgorde B..H gekoppeld aan de goal-codes uit 02_seed_reference.sql
GOAL_BY_COLUMN = {
    2: "summer_to_winter",
    3: "improve_lighting",
    4: "improve_ambiance",
    5: "replace_sky",
    6: "make_beds",
    7: "remove_object",
    8: "improve_summer",
}

STATUS_MAP = {
    "new": "new",
    "in process": "in_process",
    "qc": "qc",
    "approved": "approved",
    "denied": "denied",
    "ai rejected": "ai_rejected",
}

PRIORITY_MAP = {"low": "low", "medium": "medium", "high": "high"}

# Rijen zonder status zijn kandidaatwoningen uit de shortlist, geen opdrachten.
STATUS_WHEN_EMPTY = "backlog"

# Naamvarianten die op dezelfde persoon slaan.
EXPERT_ALIASES = {
    "Jacqueline Dubois": "Jacqueline Kunst-Dubois",
}

# Namen die geen persoon zijn maar een team.
EXPERT_TEAMS = {"Verhuur Nederland"}


# ── Hulpfuncties ─────────────────────────────────────────────────────────────

def sql_str(value):
    """Escapet een waarde voor gebruik in een SQL-string, of geeft NULL."""
    if value is None:
        return "null"
    text = str(value).strip()
    if not text:
        return "null"
    return "'" + text.replace("'", "''") + "'"


def sql_date(value, log, ctx):
    """Zet een Excel-datum om naar een SQL-datum."""
    if value is None or value == "":
        return "null"
    if isinstance(value, datetime):
        return "'" + value.date().isoformat() + "'"
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return "'" + datetime.strptime(str(value).strip(), fmt).date().isoformat() + "'"
        except ValueError:
            continue
    log.warn(f"{ctx}: datum niet herkend, overgeslagen: {value!r}")
    return "null"


def parse_photo_numbers(value, log, ctx):
    """
    Haalt fotonummers uit een cel.

    Voorkomende vormen in het bronbestand:
        3                 -> [3]
        '3, 6, 7'         -> [3, 6, 7]
        '8 en 7'          -> [8, 7]
        29.38 (float)     -> [29, 38]   Excel heeft '29,38' als decimaal gelezen
        13.1  (float)     -> [13, 1]
    """
    if value is None or value == "":
        return []

    numbers = []

    if isinstance(value, float) and not value.is_integer():
        parts = str(value).split(".")
        for part in parts:
            if part.isdigit():
                numbers.append(int(part))
        log.warn(
            f"{ctx}: cel {value!r} is door Excel als kommagetal gelezen, "
            f"geinterpreteerd als foto's {numbers}. Controleer dit handmatig."
        )
        return numbers

    if isinstance(value, (int, float)):
        return [int(value)]

    text = str(value).replace(" en ", ",").replace(";", ",").replace("/", ",")
    for token in text.split(","):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            numbers.append(int(token))
            continue
        # Vormen als '29.38' die als tekst zijn opgeslagen
        if re.fullmatch(r"\d+\.\d+", token):
            for part in token.split("."):
                if part.isdigit():
                    numbers.append(int(part))
            log.warn(f"{ctx}: token {token!r} gesplitst in losse fotonummers.")
            continue
        log.warn(f"{ctx}: token {token!r} niet herkend als fotonummer, overgeslagen.")

    return numbers


class Log:
    def __init__(self):
        self.lines = []
        self.warnings = 0

    def info(self, msg):
        self.lines.append(msg)

    def warn(self, msg):
        self.warnings += 1
        self.lines.append("LET OP  " + msg)

    def write(self, path):
        Path(path).write_text("\n".join(self.lines), encoding="utf-8")


# ── Hoofdproces ──────────────────────────────────────────────────────────────

def migrate(xlsx_path, out_path, log_path):
    log = Log()
    log.info("Migratielog VfY_FotoBewerking_Tracker.xlsx")
    log.info(f"Gedraaid op {datetime.now().isoformat(timespec='seconds')}")
    log.info(f"Bron: {xlsx_path}")
    log.info("")
    log.info("Aannames:")
    log.info("  1. Rijen zonder status zijn kandidaatwoningen en krijgen status 'backlog'.")
    log.info("  2. Kolom 'Notes QC' bevat zowel briefing vooraf als QC-feedback achteraf.")
    log.info("     De inhoud gaat daarom onbewerkt naar assignments.legacy_notes.")
    log.info("  3. Dezelfde acco_id mag meerdere keren voorkomen: dat zijn losse opdrachten.")
    log.info("  4. De kolom 'Count' wordt niet overgenomen, die volgt uit edit_items.")
    log.info("")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Tabblad {SHEET!r} niet gevonden. Aanwezig: {wb.sheetnames}")
    ws = wb[SHEET]

    experts = OrderedDict()
    editors = OrderedDict()
    assignments = []
    edit_items = []
    stats = Counter()

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        acco_id = ws.cell(row, COL["acco_id"]).value
        if acco_id is None or str(acco_id).strip() == "":
            continue

        acco_id = str(acco_id).strip()
        ctx = f"rij {row} ({acco_id})"

        raw_status = ws.cell(row, COL["status"]).value
        if raw_status is None or str(raw_status).strip() == "":
            status = STATUS_WHEN_EMPTY
            stats["backlog"] += 1
        else:
            key = str(raw_status).strip().lower()
            status = STATUS_MAP.get(key)
            if status is None:
                log.warn(f"{ctx}: onbekende status {raw_status!r}, gezet op 'new'.")
                status = "new"
            stats[status] += 1

        raw_priority = ws.cell(row, COL["priority"]).value
        priority = PRIORITY_MAP.get(str(raw_priority).strip().lower() if raw_priority else "", "low")

        expert_name = ws.cell(row, COL["rental_expert"]).value
        expert_name = str(expert_name).strip() if expert_name else None
        if expert_name:
            expert_name = EXPERT_ALIASES.get(expert_name, expert_name)
            experts.setdefault(expert_name, expert_name in EXPERT_TEAMS)

        editor_name = ws.cell(row, COL["assigned_to"]).value
        editor_name = str(editor_name).strip() if editor_name else None
        if editor_name:
            editors.setdefault(editor_name, True)

        assignments.append({
            "row": row,
            "acco_id": acco_id,
            "expert": expert_name,
            "editor": editor_name,
            "status": status,
            "priority": priority,
            "request_date": sql_date(ws.cell(row, COL["request_date"]).value, log, ctx),
            "date_assigned": sql_date(ws.cell(row, COL["date_assigned"]).value, log, ctx),
            "date_completed": sql_date(ws.cell(row, COL["date_completed"]).value, log, ctx),
            "legacy_notes": ws.cell(row, COL["notes_qc"]).value,
        })

        for col, goal_code in GOAL_BY_COLUMN.items():
            numbers = parse_photo_numbers(ws.cell(row, col).value, log, f"{ctx} / {goal_code}")
            for number in sorted(set(numbers)):
                if number <= 0:
                    log.warn(f"{ctx}: fotonummer {number} overgeslagen.")
                    continue
                edit_items.append((row, goal_code, number))

        # Controle op de oude Count-kolom
        excel_count = ws.cell(row, COL["count"]).value
        if isinstance(excel_count, (int, float)):
            unique_photos = len({n for r, g, n in edit_items if r == row})
            if int(excel_count) != unique_photos and status != "backlog":
                log.info(
                    f"{ctx}: Count in Excel was {int(excel_count)}, "
                    f"hier {unique_photos} unieke foto's. De nieuwe telling is leidend."
                )

    # ── SQL schrijven ────────────────────────────────────────────────────────
    out = []
    out.append("-- Gegenereerd door migration/migrate_xlsx.py. Niet handmatig aanpassen.")
    out.append(f"-- Bron: {Path(xlsx_path).name}")
    out.append(f"-- Gegenereerd: {datetime.now().isoformat(timespec='seconds')}")
    out.append("-- Draaien na 01_schema.sql en 02_seed_reference.sql.")
    out.append("")
    out.append("begin;")
    out.append("")

    out.append("-- Verhuurexperts")
    for name, is_team in experts.items():
        out.append(
            f"insert into rental_experts (name, is_team) values "
            f"({sql_str(name)}, {'true' if is_team else 'false'}) on conflict (name) do nothing;"
        )
    out.append("")

    out.append("-- Editors")
    for name in editors:
        out.append(
            f"insert into editors (name) values ({sql_str(name)}) on conflict (name) do nothing;"
        )
    out.append("")

    out.append("-- Opdrachten en foto's")
    out.append("-- Elke opdracht krijgt een tijdelijke sleutel via de variabele v_id.")
    out.append("do $migratie$")
    out.append("declare")
    out.append("  v_id uuid;")
    out.append("begin")

    items_by_row = {}
    for row, goal_code, number in edit_items:
        items_by_row.setdefault(row, []).append((goal_code, number))

    for a in assignments:
        expert_expr = (
            f"(select id from rental_experts where name = {sql_str(a['expert'])})"
            if a["expert"] else "null"
        )
        editor_expr = (
            f"(select id from editors where name = {sql_str(a['editor'])})"
            if a["editor"] else "null"
        )
        out.append(
            "  insert into assignments (acco_id, rental_expert_id, editor_id, status, priority, "
            "request_date, date_assigned, date_completed, legacy_notes) values ("
        )
        out.append(f"    {sql_str(a['acco_id'])}, {expert_expr}, {editor_expr},")
        out.append(f"    '{a['status']}', '{a['priority']}',")
        out.append(f"    {a['request_date']}, {a['date_assigned']}, {a['date_completed']},")
        out.append(f"    {sql_str(a['legacy_notes'])}")
        out.append("  ) returning id into v_id;")

        for goal_code, number in items_by_row.get(a["row"], []):
            out.append(
                f"  insert into edit_items (assignment_id, goal_code, photo_number) "
                f"values (v_id, '{goal_code}', {number});"
            )

    out.append("end")
    out.append("$migratie$;")
    out.append("")
    out.append("commit;")
    out.append("")

    Path(out_path).write_text("\n".join(out), encoding="utf-8")

    # ── Log afronden ─────────────────────────────────────────────────────────
    log.info("")
    log.info("Resultaat:")
    log.info(f"  opdrachten:        {len(assignments)}")
    log.info(f"  waarvan backlog:   {stats['backlog']}")
    log.info(f"  foto-goal regels:  {len(edit_items)}")
    log.info(f"  unieke foto's:     {len({(r, n) for r, g, n in edit_items})}")
    log.info(f"  verhuurexperts:    {len(experts)}")
    log.info(f"  editors:           {len(editors)}")
    log.info("  statusverdeling:   " + ", ".join(f"{k}={v}" for k, v in sorted(stats.items())))
    log.info(f"  waarschuwingen:    {log.warnings}")
    log.write(log_path)

    print(f"SQL geschreven naar {out_path}")
    print(f"Log geschreven naar {log_path} ({log.warnings} waarschuwingen)")
    return stats, len(assignments), len(edit_items)


def main():
    parser = argparse.ArgumentParser(description="Migreert de fototracker naar Supabase-SQL.")
    parser.add_argument("xlsx", help="pad naar VfY_FotoBewerking_Tracker.xlsx")
    parser.add_argument("-o", "--output", default="03_seed_data.sql")
    parser.add_argument("-l", "--log", default="migratie_log.txt")
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"Bestand niet gevonden: {args.xlsx}")

    migrate(args.xlsx, args.output, args.log)


if __name__ == "__main__":
    main()
